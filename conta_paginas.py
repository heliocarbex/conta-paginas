import argparse
import PyPDF2
from pathlib import Path
from openpyxl import Workbook
from datetime import datetime
from pathlib import _PathParents

def gera_excel_com_arquivos_e_paginas(lista_arquivos_e_paginas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Paginas"
    ws.cell(1,1).value = "Arquivo" 
    ws.cell(1,2).value = "Qtd Páginas"
    linha_atual = 2
    for [arquivo, qtd_pagina] in lista_arquivos_e_paginas:
        ws.cell(linha_atual,1).value = str(arquivo)
        ws.cell(linha_atual,2).value = qtd_pagina
        linha_atual+=1
    data_hora_atual = datetime.now().strftime("%Y-%m-%d_%Hh-%Mm-%Ss")
    arquivo_saida = f'{data_hora_atual}_lista_paginas.xlsx'

    wb.save(filename = arquivo_saida)

def conta_pagina_pdf(caminho_arquivo:str)->int:
    file = open(caminho_arquivo, 'rb')
    readpdf = PyPDF2.PdfFileReader(file)
    totalpages = readpdf.numPages
    return int(totalpages)

def junta_nomes_pastas_pai(pastas_pai:_PathParents,nivel:int):
    nomes = []
    for i in range(nivel):
        nomes.insert(0,pastas_pai[i].name)
    return '\\'.join(nomes)

def resolve_nivel_de_pastas_pai(args,i):
    if args.nivel:
        nivel = int(args.nivel)
        prefixo = junta_nomes_pastas_pai(i.parents, nivel)
    else:
        prefixo = i.parents[0].name
    return prefixo

def conta_recursivamente_paginas_pdf(args:Path):
    q_paginas = []
    somador = 0
    pasta_atual = Path(".")
    if args == '':
        pasta_arquivos = Path(".")
    else:
        pasta_arquivos = Path(".",args.pasta)
    print("Arquivos dentro da pasta\n")
    for i in pasta_arquivos.glob('**/*.pdf'):
        qtd_pagina = conta_pagina_pdf(i.resolve())
        prefixo = resolve_nivel_de_pastas_pai(args,i)
        q_paginas.append([f"{prefixo}\{i.name}", qtd_pagina])
        somador += qtd_pagina
    print(f'O total de páginas é {somador}')
    gera_excel_com_arquivos_e_paginas(q_paginas)

def main():
    parser = argparse.ArgumentParser(description='Contador de páginas de pdfs por arquivo ou por pasta')
    parser.add_argument('-a','--arquivo', help='Informe o arquivo a analisar as páginas')
    parser.add_argument('-p','--pasta', help='Informe a pasta a analisar os pdfs')
    parser.add_argument('-n','--nivel', help='Informe o nível de pastas pai para imprimir')

    args = parser.parse_args()

    if args.arquivo:
        qtd_pagina = conta_pagina_pdf(Path(args.arquivo).resolve())
        print(f"A quantidade de páginas de {args.arquivo} é {qtd_pagina}")
    elif args.pasta:
        conta_recursivamente_paginas_pdf(args)
    else:
        conta_recursivamente_paginas_pdf("")

if __name__ == "__main__":
    main()